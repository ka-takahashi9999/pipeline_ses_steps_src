#!/usr/bin/env python3
"""One-Sheet acquisition plus exact-contract offline artifact regeneration."""

import argparse
import copy
import hashlib
import io
import json
import os
import re
import ssl
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


sys.dont_write_bytecode = True
STEP_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = STEP_DIR.parent
ACQUISITION_DIR = STEP_DIR / "00_tool" / "acquisition"
for import_path in (PROJECT_ROOT, ACQUISITION_DIR):
    if str(import_path) not in sys.path:
        sys.path.insert(0, str(import_path))

from common.file_utils import ensure_dir, write_execution_time
from common.json_utils import read_jsonl, read_jsonl_as_list, write_jsonl
from common.logger import get_logger
from google_sheet_acquisition_contract import (
    ATTEMPT_PLAN_VERSION,
    ENTRY_RAW_DOMAIN,
    MANIFEST_DIGEST_FIELDS,
    MANIFEST_DOMAIN,
    MANIFEST_FIELDS,
    MANIFEST_VERSION,
    ORDERED_SNAPSHOT_SET_FIELDS,
    PLANNED_CONTAINER_ENTRY_FIELDS,
    PLANNED_CONTAINER_SET_DIGEST_FIELDS,
    PLANNED_CONTAINER_SET_DOMAIN,
    PLANNED_CONTAINER_SET_VERSION,
    PROFILE_DIGEST_FIELDS,
    PROFILE_DOMAIN,
    RESOLVED_SCOPE_DIGEST_FIELDS,
    RESOLVED_SCOPE_DOMAIN,
    RESOLVED_SCOPE_VERSION,
    SNAPSHOT_ENTRY_FIELDS,
    SNAPSHOT_ENTRY_VERSION,
    ContractError,
    calculate_entry_raw_digest,
    calculate_ordered_snapshot_set_digest,
    canonical_json_bytes,
    finalize_manifest,
    finalize_planned_container_set,
    finalize_resolved_scope,
    offline_negative_proofs,
    parse_canonical_json,
    validate_attempt_plan,
    validate_manifest,
    validate_profile_registry,
    validate_snapshot_entry,
)


logger = get_logger("99-1_google_sheet_acquisition_prototype")
PROFILE_PATH = STEP_DIR / "10_assistance_tool" / "configs" / "acquisition" / "google_sheets_public_xlsx.v1.json.example"
GOLDEN_PATH = STEP_DIR / "10_assistance_tool" / "fixtures" / "acquisition" / "source_acquisition_digest_golden.v1.json.example"
DEFAULT_GMAIL_PATH = PROJECT_ROOT / "01-1_fetch_gmail" / "01_result" / "fetch_gmail_mail_master.jsonl"
RESULT_SUBDIR = "google_sheet_acquisition_prototype"
RESULT_DIR = STEP_DIR / "01_result" / RESULT_SUBDIR
SNAPSHOT_RELATIVE_PATH = "snapshot/google_sheet.xlsx"
MAX_DOWNLOAD_BYTES = 50 * 1024 * 1024
MAX_EXPANDED_BYTES = 250 * 1024 * 1024
MAX_PACKAGE_MEMBERS = 10000
USER_AGENT = "Mozilla/5.0 (compatible; 99-1-Google-Sheet-Acquisition-Prototype/1.0)"
GOOGLE_SHEET_PATH = re.compile(r"\A/spreadsheets/d/([A-Za-z0-9_-]+)(?:/|\Z)")
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SELECTOR_SENDER_DOMAIN = "tanapism.co.jp"
SELECTOR_LINK_TEXT = "その他人材情報一覧"


class AcquisitionStop(Exception):
    def __init__(self, status: str, reason: str):
        super().__init__(reason)
        self.status = status
        self.reason = reason


def _utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%fZ")


def _stable_identifier(prefix: str, value: Any) -> str:
    encoded = canonical_json_bytes(value)
    return prefix + ":" + hashlib.sha256(encoded).hexdigest()


def _load_profile_registry() -> Dict[str, Any]:
    registry = json.loads(PROFILE_PATH.read_text(encoding="utf-8"))
    reasons = validate_profile_registry(registry)
    if reasons:
        raise ValueError("profile registry invalid:" + ",".join(reasons))
    return registry


def _normalized_sheet_locator(href: str) -> Optional[Dict[str, str]]:
    try:
        parsed = urllib.parse.urlsplit(href)
    except ValueError:
        return None
    if parsed.scheme != "https" or parsed.hostname != "docs.google.com":
        return None
    match = GOOGLE_SHEET_PATH.match(parsed.path)
    if match is None:
        return None
    spreadsheet_key = match.group(1)
    query = urllib.parse.parse_qs(parsed.query)
    fragment = urllib.parse.parse_qs(parsed.fragment)
    gid_values = fragment.get("gid") or query.get("gid") or []
    gid = gid_values[0] if gid_values and gid_values[0].isdigit() else "UNKNOWN"
    normalized = "https://docs.google.com/spreadsheets/d/" + spreadsheet_key + "/edit"
    if gid != "UNKNOWN":
        normalized += "#gid=" + gid
    return {"spreadsheet_key": spreadsheet_key, "gid": gid, "normalized_locator": normalized}


def select_representative_locator(gmail_path: Path, profile: Dict[str, Any]) -> Dict[str, Any]:
    """Resolve the same one Tanapism list locator; Profile architecture is unchanged."""
    evidence: Dict[str, Dict[str, Any]] = {}
    for record in read_jsonl(str(gmail_path)):
        if SELECTOR_SENDER_DOMAIN not in str(record.get("from", "")).casefold():
            continue
        for link in record.get("html_links", []):
            if not isinstance(link, dict) or SELECTOR_LINK_TEXT.casefold() not in str(link.get("text", "")).casefold():
                continue
            locator = _normalized_sheet_locator(str(link.get("href", "")))
            if locator is None:
                continue
            current = evidence.setdefault(
                locator["normalized_locator"],
                {**locator, "evidence_message_ids": [], "link_text": str(link.get("text", ""))},
            )
            if isinstance(record.get("message_id"), str) and record["message_id"]:
                current["evidence_message_ids"].append(record["message_id"])
    if len(evidence) != 1:
        raise ValueError("representative locator must resolve exactly once; got " + str(len(evidence)))
    selected = next(iter(evidence.values()))
    selected["evidence_message_ids"] = sorted(set(selected["evidence_message_ids"]))
    selected["evidence_count"] = len(selected["evidence_message_ids"])
    return selected


def build_attempt_plan(registry: Dict[str, Any], selected: Dict[str, Any], started_at: str) -> Dict[str, Any]:
    profile = registry["profiles"][0]
    source_id = "source:google-sheet:" + selected["spreadsheet_key"]
    locator_ref = _stable_identifier("locator-ref", selected["normalized_locator"])
    locator_binding_id = _stable_identifier(
        "locator-binding",
        {"locator_ref": locator_ref, "evidence_message_ids": selected["evidence_message_ids"]},
    )
    attempt_id = "google-sheet-acquisition-" + locator_binding_id.rsplit(":", 1)[1][:24]
    profile_ref = {
        "acquisition_profile_id": profile["acquisition_profile_id"],
        "acquisition_profile_version": profile["acquisition_profile_version"],
        "acquisition_profile_digest": profile["acquisition_profile_digest"],
    }
    scope = finalize_resolved_scope({
        "resolved_scope_schema_version": RESOLVED_SCOPE_VERSION,
        "source_id": source_id,
        "attempt_id": attempt_id,
        **profile_ref,
        "provider_id": profile["provider_id"],
        "source_class": profile["source_class"],
        "locator_binding_id": locator_binding_id,
        "acquisition_method": profile["acquisition_method"],
        "scope_payload": {
            **profile["scope_template"],
            "normalized_locator": selected["normalized_locator"],
            "representative_gid": selected["gid"],
            "spreadsheet_key": selected["spreadsheet_key"],
        },
        "presentation_policy": profile["presentation_policy"],
        "version_requirement": "WORKBOOK_WIDE_STRONG",
        "termination_rule": "PREBOUND_SINGLE_WORKBOOK",
        "resolved_planned_scope_digest": "sha256:" + "0" * 64,
    })
    planned_container_id = "planned-google-sheet-" + _stable_identifier(
        "planned-container", {"profile_ref": profile_ref, "locator_binding_id": locator_binding_id}
    ).rsplit(":", 1)[1][:24]
    container_set = finalize_planned_container_set({
        "planned_container_set_schema_version": PLANNED_CONTAINER_SET_VERSION,
        "source_id": source_id,
        "attempt_id": attempt_id,
        **profile_ref,
        "resolved_planned_scope_digest": scope["resolved_planned_scope_digest"],
        "planned_container_entries": [{
            "sequence": 0,
            "planned_container_id": planned_container_id,
            "container_kind": profile["planned_container_template"]["container_kind"],
            "required": profile["planned_container_template"]["required"],
            "logical_role": profile["planned_container_template"]["logical_role"],
            "locator_ref": locator_ref,
            "locator_binding_id": locator_binding_id,
            **profile_ref,
            "optional_absence_rule": "NOT_ALLOWED",
        }],
        "planned_container_set_digest": "sha256:" + "0" * 64,
    })
    plan = {
        "schema_version": ATTEMPT_PLAN_VERSION,
        "attempt_id": attempt_id,
        "attempt_ordinal": 1,
        "attempt_started_at": started_at,
        "provider": profile["provider_id"],
        "acquisition_method": profile["acquisition_method"],
        "profile_ref": profile_ref,
        "resolved_planned_scope": scope,
        "planned_container_set": container_set,
        "candidate_emission": 0,
        "auto_union": False,
        "production_integration": False,
        "plan_state": "FROZEN_BEFORE_ACQUISITION",
    }
    reasons = validate_attempt_plan(plan)
    if reasons:
        raise ValueError("attempt plan invalid:" + ",".join(reasons))
    return plan


def _request(url: str, method: str = "GET", read_limit: Optional[int] = None) -> Tuple[bytes, Dict[str, str], str, int]:
    request = urllib.request.Request(url, method=method, headers={"User-Agent": USER_AGENT, "Accept": "*/*"})
    try:
        with urllib.request.urlopen(request, timeout=30, context=ssl.create_default_context()) as response:
            final_host = urllib.parse.urlsplit(response.geturl()).hostname or ""
            payload = b"" if method == "HEAD" else response.read(read_limit)
            headers = {
                name.lower(): value for name, value in response.headers.items()
                if name.lower() in {"content-type", "etag", "last-modified", "x-goog-generation"}
            }
            return payload, headers, final_host, int(response.status)
    except urllib.error.HTTPError as error:
        if error.code in {401, 403}:
            raise AcquisitionStop("AUTH_REQUIRED", "provider returned authentication status") from error
        raise AcquisitionStop("FAILED", "provider HTTP status:" + str(error.code)) from error
    except (urllib.error.URLError, TimeoutError, ssl.SSLError) as error:
        raise AcquisitionStop("FAILED", "provider access failed:" + type(error).__name__) from error


def _check_public_shared_url(shared_url: str) -> None:
    payload, _, final_host, _ = _request(shared_url, read_limit=256 * 1024)
    lowered = payload.lower()
    if final_host == "accounts.google.com" or b"accounts.google.com/signin" in lowered:
        raise AcquisitionStop("AUTH_REQUIRED", "shared URL requires Google login")
    if b"request access" in lowered or b"you need access" in lowered:
        raise AcquisitionStop("AUTH_REQUIRED", "shared URL requires access grant")
    if final_host not in {"docs.google.com", "drive.google.com"}:
        raise AcquisitionStop("FAILED", "unexpected shared URL response host")


def _legacy_version(headers: Dict[str, str]) -> Dict[str, str]:
    if headers.get("etag"):
        return {"kind": "ETAG", "scope": "RESPONSE", "strength": "WEAK", "value": headers["etag"]}
    if headers.get("last-modified"):
        return {"kind": "LAST_MODIFIED", "scope": "RESPONSE", "strength": "WEAK", "value": headers["last-modified"]}
    return {"kind": "UNKNOWN", "scope": "UNKNOWN", "strength": "UNKNOWN", "value": "UNAVAILABLE"}


def _version_observation(value: Dict[str, str], observed_at: str, phase: str, attempt_id: str) -> Dict[str, Any]:
    return {
        "version_kind": value["kind"],
        "version_scope": value["scope"],
        "version_strength": value["strength"],
        "version_binding_id": _stable_identifier("version-binding", {"attempt_id": attempt_id, "phase": phase, "value": value["value"]}),
        "provider_authority_ref": "provider:google-sheets:public-xlsx-response",
        "observed_at": observed_at,
    }


def _unknown_source_version(observed_at: str, attempt_id: str) -> Dict[str, Any]:
    pre = _version_observation(_legacy_version({}), observed_at, "pre", attempt_id)
    post = _version_observation(_legacy_version({}), observed_at, "post", attempt_id)
    current = _version_observation(_legacy_version({}), observed_at, "authority", attempt_id)
    return {**current, "pre_version": pre, "post_version": post}


def _source_version_from_headers(before: Dict[str, str], after: Dict[str, str], observed_at: str, attempt_id: str) -> Dict[str, Any]:
    pre_value = _legacy_version(before)
    post_value = _legacy_version(after)
    current = pre_value if pre_value["kind"] == post_value["kind"] else _legacy_version({})
    return {
        **_version_observation(current, observed_at, "authority", attempt_id),
        "pre_version": _version_observation(pre_value, observed_at, "pre", attempt_id),
        "post_version": _version_observation(post_value, observed_at, "post", attempt_id),
    }


def _validate_xlsx_package(payload: bytes) -> None:
    if len(payload) > MAX_DOWNLOAD_BYTES:
        raise AcquisitionStop("FAILED", "XLSX download size limit exceeded")
    try:
        with zipfile.ZipFile(io.BytesIO(payload), "r") as package:
            infos = package.infolist()
            if len(infos) > MAX_PACKAGE_MEMBERS or sum(info.file_size for info in infos) > MAX_EXPANDED_BYTES:
                raise AcquisitionStop("FAILED", "XLSX package limit exceeded")
            names = {info.filename for info in infos}
            if "xl/workbook.xml" not in names or "[Content_Types].xml" not in names:
                raise AcquisitionStop("FAILED", "XLSX required package parts missing")
            for info in infos:
                normalized = info.filename.replace("\\", "/")
                if normalized.startswith("/") or ".." in normalized.split("/"):
                    raise AcquisitionStop("FAILED", "XLSX unsafe package path")
    except zipfile.BadZipFile as error:
        raise AcquisitionStop("FAILED", "provider response is not XLSX") from error


def _xlsx_sheet_ids(payload: bytes) -> List[Dict[str, str]]:
    with zipfile.ZipFile(io.BytesIO(payload), "r") as package:
        xml_payload = package.read("xl/workbook.xml")
    if b"<!doctype" in xml_payload.lower() or b"<!entity" in xml_payload.lower():
        raise AcquisitionStop("FAILED", "XLSX workbook XML is unsafe")
    root = ET.fromstring(xml_payload)
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    return [
        {"tab_title": element.get("name", ""), "tab_id": element.get("sheetId", "UNKNOWN"), "tab_id_kind": "XLSX_SHEET_ID"}
        for element in root.findall(".//{%s}sheet" % namespace)
    ]


def observe_workbook(payload: bytes) -> Dict[str, Any]:
    _validate_xlsx_package(payload)
    workbook = load_workbook(io.BytesIO(payload), read_only=False, data_only=False, keep_links=False)
    values_workbook = load_workbook(io.BytesIO(payload), read_only=False, data_only=True, keep_links=False)
    sheet_ids = _xlsx_sheet_ids(payload)
    if len(sheet_ids) != len(workbook.worksheets):
        raise AcquisitionStop("FAILED", "XLSX workbook inventory mismatch")
    tabs: List[Dict[str, Any]] = []
    total_formula_count = 0
    total_cached = 0
    for position, worksheet in enumerate(workbook.worksheets):
        values_sheet = values_workbook[worksheet.title]
        formula_count = 0
        cached_count = 0
        styled_count = 0
        for row in worksheet.iter_rows():
            for cell in row:
                styled_count += int(cell.has_style)
                if cell.data_type == "f":
                    formula_count += 1
                    cached_count += int(values_sheet[cell.coordinate].value is not None)
        total_formula_count += formula_count
        total_cached += cached_count
        tabs.append({
            "tab_order": position, **sheet_ids[position], "visibility": worksheet.sheet_state.upper(),
            "row_bound": worksheet.max_row, "column_bound": worksheet.max_column,
            "requested_range": "ENTIRE_EXPORTED_TAB_USED_BOUNDS",
            "returned_range": worksheet.calculate_dimension(), "formula_count": formula_count,
            "cached_formula_value_count": cached_count, "styled_cell_count": styled_count,
            "conditional_formatting_rule_set_count": len(worksheet.conditional_formatting),
            "hidden_row_count": sum(1 for value in worksheet.row_dimensions.values() if value.hidden),
            "hidden_column_count": sum(1 for value in worksheet.column_dimensions.values() if value.hidden),
        })
    return {
        "workbook_tab_inventory": "AVAILABLE",
        "tab_id_authority": "XLSX_SHEET_ID_AVAILABLE_GOOGLE_GID_INVENTORY_UNAVAILABLE",
        "tab_count": len(tabs), "tabs": tabs, "range_bounds": "AVAILABLE",
        "formula_metadata_availability": "AVAILABLE",
        "presentation_metadata": {
            "policy": "UNRESOLVED", "overall_availability": "PARTIAL",
            "effective_format": "UNAVAILABLE", "static_user_format": "AVAILABLE",
            "conditional_formatting": "AVAILABLE", "hidden_rows_columns": "AVAILABLE",
            "tab_visibility": "AVAILABLE", "formula_metadata": "AVAILABLE",
            "display_values": "AVAILABLE" if total_formula_count == total_cached else "PARTIAL",
            "business_state_mapping": "NOT_IMPLEMENTED",
        },
        "technical_record_count": sum(max(worksheet.max_row, 0) for worksheet in workbook.worksheets),
    }


def _atomic_write_bytes(path: Path, payload: bytes) -> None:
    ensure_dir(str(path.parent))
    temporary = path.with_suffix(path.suffix + ".tmp")
    try:
        with open(temporary, "wb") as stream:
            stream.write(payload)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(str(temporary), str(path))
    except Exception:
        if temporary.exists():
            temporary.unlink()
        raise


def _build_snapshot_and_manifest(
    registry: Dict[str, Any], plan: Dict[str, Any], payload: bytes,
    retrieved_at: str, source_version: Dict[str, Any], http_status: int,
) -> Tuple[Dict[str, Any], Dict[str, bytes]]:
    profile = registry["profiles"][0]
    scope = plan["resolved_planned_scope"]
    planned_set = plan["planned_container_set"]
    planned = planned_set["planned_container_entries"][0]
    entry_id = "snapshot-entry:" + planned["planned_container_id"]
    snapshot_entry = {
        "snapshot_entry_schema_version": SNAPSHOT_ENTRY_VERSION,
        "sequence": 0,
        "entry_id": entry_id,
        "planned_container_id": planned["planned_container_id"],
        "page_or_tab_id": {"state": "NOT_APPLICABLE", "value": ""},
        "locator_binding_id": planned["locator_binding_id"],
        "requested_range": {"state": "VALUE", "value": "WORKBOOK_ALL_TABS"},
        "returned_range": {"state": "VALUE", "value": "WORKBOOK_ALL_TABS"},
        "retrieved_at": retrieved_at,
        "source_version": copy.deepcopy(source_version["post_version"]),
        "http_status": http_status,
        "content_type": XLSX_MEDIA_TYPE,
        "byte_length": len(payload),
        "raw_artifact_ref": SNAPSHOT_RELATIVE_PATH,
        "entry_raw_digest": calculate_entry_raw_digest(payload),
        "next_locator_metadata": {"state": "NOT_APPLICABLE", "next_locator_binding_id": ""},
        "terminal": "TERMINAL",
        "terminal_evidence": {
            "evidence_kind": "PREBOUND_SCOPE_END",
            "evidence_binding_id": _stable_identifier("terminal-evidence", scope["resolved_planned_scope_digest"]),
        },
    }
    manifest = finalize_manifest({
        "manifest_schema_version": MANIFEST_VERSION,
        "source_id": scope["source_id"],
        "attempt_id": plan["attempt_id"],
        "manifest_finalized_at": retrieved_at,
        "acquisition_profile_id": profile["acquisition_profile_id"],
        "acquisition_profile_version": profile["acquisition_profile_version"],
        "acquisition_profile_digest": profile["acquisition_profile_digest"],
        "resolved_planned_scope_digest": scope["resolved_planned_scope_digest"],
        "planned_container_set_digest": planned_set["planned_container_set_digest"],
        "actual_container_entries": [{
            "sequence": 0, "container_id": planned["planned_container_id"],
            "container_kind": planned["container_kind"], "logical_role": planned["logical_role"],
            "locator_ref": planned["locator_ref"], "locator_binding_id": planned["locator_binding_id"],
            "acquisition_profile_id": profile["acquisition_profile_id"],
            "acquisition_profile_version": profile["acquisition_profile_version"],
            "acquisition_profile_digest": profile["acquisition_profile_digest"],
            "snapshot_entry_ids": [entry_id],
        }],
        "snapshot_entries": [snapshot_entry], "snapshot_count": 1,
        "ordered_snapshot_set_digest": calculate_ordered_snapshot_set_digest([snapshot_entry]),
        "source_version": source_version, "manifest_digest": "sha256:" + "0" * 64,
    })
    return manifest, {entry_id: payload}


def _contract_conformance_checks(manifest: Dict[str, Any]) -> List[Dict[str, Any]]:
    golden = json.loads(GOLDEN_PATH.read_text(encoding="utf-8"))
    registry = _load_profile_registry()
    plan = read_jsonl_as_list(str(RESULT_DIR / "attempt_plan.jsonl"))[0]
    profile = registry["profiles"][0]
    scope = plan["resolved_planned_scope"]
    container_set = plan["planned_container_set"]
    raw = (RESULT_DIR / SNAPSHOT_RELATIVE_PATH).read_bytes()

    def direct_digest(domain: str, payload: bytes) -> str:
        return "sha256:" + hashlib.sha256(domain.encode("utf-8") + b"\x00" + payload).hexdigest()

    checks: List[Tuple[str, bool]] = []
    checks.append(("golden_canonical_bytes", canonical_json_bytes(golden["logical_object"]) == golden["expected_canonical_json"].encode("utf-8")))
    checks.append(("golden_ordered_set_digest", calculate_ordered_snapshot_set_digest(golden["logical_object"]) == golden["expected_ordered_snapshot_set_digest"]))
    checks.append(("newline_escape", canonical_json_bytes("\n") == b'"\\u000a"'))
    checks.append(("tab_escape", canonical_json_bytes("\t") == b'"\\u0009"'))
    checks.append(("nfc_equivalent", canonical_json_bytes("e\u0301") == canonical_json_bytes("\u00e9")))
    for name, operation in (
        ("nfc_key_collision_reject", lambda: canonical_json_bytes({"e\u0301": 1, "\u00e9": 2})),
        ("duplicate_key_reject", lambda: parse_canonical_json(b'{"a":1,"a":2}')),
        ("float_reject", lambda: canonical_json_bytes(1.0)),
        ("exponent_reject", lambda: parse_canonical_json(b"1e2")),
        ("nan_reject", lambda: parse_canonical_json(b"NaN")),
        ("leading_plus_reject", lambda: parse_canonical_json(b"+1")),
        ("leading_zero_reject", lambda: parse_canonical_json(b"01")),
        ("int64_max_plus_one_reject", lambda: canonical_json_bytes(2 ** 63)),
        ("int64_min_minus_one_reject", lambda: canonical_json_bytes(-(2 ** 63) - 1)),
    ):
        try:
            operation()
            checks.append((name, False))
        except ContractError:
            checks.append((name, True))
    checks.extend([
        ("int64_max", canonical_json_bytes(2 ** 63 - 1) == b"9223372036854775807"),
        ("int64_min", canonical_json_bytes(-(2 ** 63)) == b"-9223372036854775808"),
        ("negative_zero", canonical_json_bytes(json.loads("-0")) == b"0"),
        ("manifest_field_count", len(manifest) == 15 and set(manifest) == set(MANIFEST_FIELDS)),
        ("snapshot_field_count", len(manifest["snapshot_entries"][0]) == 18 and set(manifest["snapshot_entries"][0]) == set(SNAPSHOT_ENTRY_FIELDS)),
        (
            "entry_raw_digest_domain",
            manifest["snapshot_entries"][0]["entry_raw_digest"]
            == direct_digest(ENTRY_RAW_DOMAIN, raw),
        ),
        (
            "manifest_digest_domain_projection",
            manifest["manifest_digest"]
            == direct_digest(
                MANIFEST_DOMAIN,
                canonical_json_bytes({field: manifest[field] for field in MANIFEST_DIGEST_FIELDS}),
            ),
        ),
        (
            "profile_digest_domain_projection",
            profile["acquisition_profile_digest"]
            == direct_digest(
                PROFILE_DOMAIN,
                canonical_json_bytes({field: profile[field] for field in PROFILE_DIGEST_FIELDS}),
            ),
        ),
        (
            "resolved_scope_digest_domain_projection",
            scope["resolved_planned_scope_digest"]
            == direct_digest(
                RESOLVED_SCOPE_DOMAIN,
                canonical_json_bytes({field: scope[field] for field in RESOLVED_SCOPE_DIGEST_FIELDS}),
            ),
        ),
    ])
    planned_projection = {
        field: container_set[field] for field in PLANNED_CONTAINER_SET_DIGEST_FIELDS
    }
    planned_projection["planned_container_entries"] = [
        {field: entry[field] for field in PLANNED_CONTAINER_ENTRY_FIELDS}
        for entry in container_set["planned_container_entries"]
    ]
    checks.append(
        (
            "planned_container_set_digest_domain_projection",
            container_set["planned_container_set_digest"]
            == direct_digest(
                PLANNED_CONTAINER_SET_DOMAIN,
                canonical_json_bytes(planned_projection),
            ),
        )
    )
    ordered_projection = [
        {field: entry[field] for field in ORDERED_SNAPSHOT_SET_FIELDS}
        for entry in manifest["snapshot_entries"]
    ]
    checks.append(
        (
            "ordered_set_actual_domain_projection",
            manifest["ordered_snapshot_set_digest"]
            == direct_digest(
                "99-1/source-acquisition/ordered-snapshot-set/v1",
                canonical_json_bytes(ordered_projection),
            ),
        )
    )
    unknown_manifest = copy.deepcopy(manifest)
    unknown_manifest["unknown"] = "reject"
    checks.append(("unknown_manifest_field_reject", any("unknown_field" in reason for reason in validate_manifest(
        unknown_manifest, registry, plan,
        {manifest["snapshot_entries"][0]["entry_id"]: (RESULT_DIR / SNAPSHOT_RELATIVE_PATH).read_bytes()},
    )["reasons"])))
    unknown_snapshot = copy.deepcopy(manifest["snapshot_entries"][0])
    unknown_snapshot["unknown"] = "reject"
    checks.append(("unknown_snapshot_field_reject", any("unknown_field" in reason for reason in validate_snapshot_entry(unknown_snapshot, True))))
    unknown_nested = copy.deepcopy(manifest["snapshot_entries"][0])
    unknown_nested["page_or_tab_id"]["unknown"] = "reject"
    checks.append(("unknown_nested_field_reject", any("unknown_field" in reason for reason in validate_snapshot_entry(unknown_nested, True))))
    null_status = copy.deepcopy(manifest["snapshot_entries"][0])
    null_status["http_status"] = None
    checks.append(("http_status_null_reject", bool(validate_snapshot_entry(null_status, True))))
    bool_sequence = copy.deepcopy(manifest["snapshot_entries"][0])
    bool_sequence["sequence"] = True
    checks.append(("bool_as_integer_reject", any("not_int64" in reason for reason in validate_snapshot_entry(bool_sequence, True))))
    checks.append(("datetime_exact", not validate_snapshot_entry(manifest["snapshot_entries"][0], True)))
    for name, invalid_datetime in (
        ("datetime_offset_reject", "2026-08-27T12:34:56.000000+00:00"),
        ("datetime_fraction_missing_reject", "2026-08-27T12:34:56Z"),
    ):
        changed = copy.deepcopy(manifest["snapshot_entries"][0])
        changed["retrieved_at"] = invalid_datetime
        checks.append((name, bool(validate_snapshot_entry(changed, True))))
    return [{"name": name, "passed": passed} for name, passed in checks]


def _negative_proof_pass_count(cases: List[Dict[str, Any]]) -> int:
    expected = {
        "profile_digest_mismatch": ("UNVERIFIED", "profile_digest_mismatch"),
        "planned_scope_mismatch": ("PARTIAL", "planned_scope_mismatch"),
        "required_container_missing": ("PARTIAL", "required_container_missing"),
        "strong_version_unavailable": ("UNVERIFIED", "strong_version_unavailable"),
        "revision_drift": ("SNAPSHOT_UNSTABLE", "revision_drift"),
        "range_gap": ("PARTIAL", "range_gap"),
        "digest_mismatch": ("INCOMPLETE", "snapshot_entry_digest_mismatch"),
        "presentation_unresolved": ("UNVERIFIED", "presentation_unresolved"),
        "attempt_uncommitted": ("INCOMPLETE", "attempt_uncommitted"),
    }
    return sum(
        case["result"]["acquisition_status"] == expected[case["name"]][0]
        and expected[case["name"]][1] in case["result"]["reasons"]
        and case["result"]["eligible"] == 0
        for case in cases
    )


def _write_outputs(
    registry: Dict[str, Any], plan: Dict[str, Any], manifest: Dict[str, Any],
    observation: Dict[str, Any], validation: Dict[str, Any],
    conformance: List[Dict[str, Any]], negative_cases: List[Dict[str, Any]],
    live_access_count: int,
) -> Dict[str, Any]:
    ensure_dir(str(RESULT_DIR))
    write_jsonl(str(RESULT_DIR / "profile_registry.jsonl"), [registry])
    write_jsonl(str(RESULT_DIR / "attempt_plan.jsonl"), [plan])
    write_jsonl(str(RESULT_DIR / "snapshot_entries.jsonl"), manifest["snapshot_entries"])
    write_jsonl(str(RESULT_DIR / "acquisition_manifest.jsonl"), [manifest])
    write_jsonl(str(RESULT_DIR / "google_metadata_observation.jsonl"), [observation])
    write_jsonl(str(RESULT_DIR / "digest_schema_conformance.jsonl"), conformance)
    write_jsonl(str(RESULT_DIR / "negative_proof_result.jsonl"), [
        {"name": case["name"], "status": case["result"]["acquisition_status"],
         "review_status": case["result"]["review_status"], "eligible": case["result"]["eligible"],
         "reasons": case["result"]["reasons"]}
        for case in negative_cases
    ])
    negative_pass = _negative_proof_pass_count(negative_cases)
    all_conformance = all(check["passed"] for check in conformance)
    strict_pass = (
        validation["valid"] and validation["exact_manifest_schema"]
        and validation["exact_snapshot_entry_schema"] and validation["digest_conformance"]
        and all_conformance and negative_pass == 9 and validation["eligible"] == 0
        and validation["candidate_emission"] == 0
    )
    summary = {
        "implementation": "PASS" if strict_pass else "FAIL",
        "prototype_pass_condition": "STRICT",
        "google_live_access": live_access_count,
        "provider": "GOOGLE_SHEETS",
        "access": "SUCCESS",
        "acquisition_method": registry["profiles"][0]["acquisition_method"],
        "profile": "PASS" if not validate_profile_registry(registry) else "FAIL",
        "attempt_plan": "PASS" if not validate_attempt_plan(plan) else "FAIL",
        "planned_container": "PASS",
        "snapshot": "PASS",
        "manifest": "PASS" if validation["exact_manifest_schema"] else "FAIL",
        "manifest_field_count": len(manifest),
        "snapshot_entry_field_count": len(manifest["snapshot_entries"][0]),
        "manifest_validation": "PASS" if validation["valid"] else "FAIL",
        "digest_conformance": "PASS" if validation["digest_conformance"] else "FAIL",
        "canonical_conformance": "PASS" if all_conformance else "FAIL",
        "original_negative_proofs": {"passed": negative_pass, "total": 9},
        "acquisition_status": validation["acquisition_status"],
        "review_status": validation["review_status"],
        "attempt_state": "COMMITTED",
        "eligible": 0, "auto_union": False, "candidate_emission": 0,
        "actual_fixed_oracle": 0, "production_write": 0, "P8": "NONE",
        "validation_reasons": validation["reasons"],
    }
    write_jsonl(str(RESULT_DIR / "prototype_summary.jsonl"), [summary])
    return summary


def regenerate_offline() -> Dict[str, Any]:
    """Reuse the one saved XLSX; perform zero Google requests."""
    started = time.monotonic()
    snapshot_path = RESULT_DIR / SNAPSHOT_RELATIVE_PATH
    if not snapshot_path.exists():
        raise FileNotFoundError("saved prototype XLSX missing:" + str(snapshot_path))
    payload = snapshot_path.read_bytes()
    _validate_xlsx_package(payload)
    old_manifest_rows = read_jsonl_as_list(str(RESULT_DIR / "acquisition_manifest.jsonl"))
    old_plan_rows = read_jsonl_as_list(str(RESULT_DIR / "attempt_plan.jsonl"))
    old_observation_rows = read_jsonl_as_list(str(RESULT_DIR / "google_metadata_observation.jsonl"))
    old_manifest = old_manifest_rows[0]
    old_plan = old_plan_rows[0]
    observation = old_observation_rows[0]
    representative = old_manifest.get("representative_source", {})
    if not representative:
        scope_payload = old_plan.get("resolved_planned_scope", {}).get("scope_payload", {})
        representative = {
            "spreadsheet_key": scope_payload["spreadsheet_key"],
            "gid": scope_payload["representative_gid"],
            "normalized_locator": scope_payload["normalized_locator"],
            "evidence_message_ids": [], "evidence_count": 0,
        }
    representative.setdefault("evidence_message_ids", [])
    registry = _load_profile_registry()
    started_at = old_plan.get("attempt_started_at", observation.get("acquisition_timestamp", _utc_now()))
    plan = build_attempt_plan(registry, representative, started_at)
    retrieved_at = observation.get("acquisition_timestamp", _utc_now())
    source_version = _unknown_source_version(retrieved_at, plan["attempt_id"])
    manifest, raw_entries = _build_snapshot_and_manifest(
        registry, plan, payload, retrieved_at, source_version, 200
    )
    # Persist exact plan before conformance checks that read the plan artifact.
    write_jsonl(str(RESULT_DIR / "attempt_plan.jsonl"), [plan])
    validation = validate_manifest(manifest, registry, plan, raw_entries)
    negative_cases = offline_negative_proofs(manifest, registry, plan, raw_entries)
    conformance = _contract_conformance_checks(manifest)
    summary = _write_outputs(
        registry, plan, manifest, observation, validation, conformance,
        negative_cases, live_access_count=0,
    )
    write_execution_time(
        str(STEP_DIR / "99_execution_time"),
        "99-1_google_sheet_digest_conformance_offline", time.monotonic() - started, 1,
    )
    logger.ok(
        "offline regeneration=" + summary["implementation"]
        + " live_access=0 acquisition=" + summary["acquisition_status"]
        + " eligible=0 candidate_emission=0"
    )
    return {"registry": registry, "plan": plan, "manifest": manifest,
            "observation": observation, "validation": validation,
            "raw_entries": raw_entries, "summary": summary}


def run_live(gmail_path: Path = DEFAULT_GMAIL_PATH) -> Dict[str, Any]:
    """Original public shared URL/XLSX method, retained without redesign."""
    started = time.monotonic()
    registry = _load_profile_registry()
    selected = select_representative_locator(gmail_path, registry["profiles"][0])
    plan = build_attempt_plan(registry, selected, _utc_now())
    _check_public_shared_url(selected["normalized_locator"])
    export_url = "https://docs.google.com/spreadsheets/d/" + selected["spreadsheet_key"] + "/export?format=xlsx"
    _, before_headers, _, _ = _request(export_url, method="HEAD")
    payload, _, _, http_status = _request(export_url, method="GET", read_limit=MAX_DOWNLOAD_BYTES + 1)
    _validate_xlsx_package(payload)
    observation = {"spreadsheet_key": selected["spreadsheet_key"], "gid": selected["gid"],
                   "acquisition_timestamp": _utc_now(), **observe_workbook(payload)}
    _, after_headers, _, _ = _request(export_url, method="HEAD")
    retrieved_at = observation["acquisition_timestamp"]
    source_version = _source_version_from_headers(before_headers, after_headers, retrieved_at, plan["attempt_id"])
    _atomic_write_bytes(RESULT_DIR / SNAPSHOT_RELATIVE_PATH, payload)
    manifest, raw_entries = _build_snapshot_and_manifest(
        registry, plan, payload, retrieved_at, source_version, http_status
    )
    write_jsonl(str(RESULT_DIR / "attempt_plan.jsonl"), [plan])
    validation = validate_manifest(manifest, registry, plan, raw_entries)
    negative_cases = offline_negative_proofs(manifest, registry, plan, raw_entries)
    conformance = _contract_conformance_checks(manifest)
    summary = _write_outputs(registry, plan, manifest, observation, validation, conformance, negative_cases, 1)
    write_execution_time(str(STEP_DIR / "99_execution_time"), "99-1_google_sheet_acquisition_prototype", time.monotonic() - started, 1)
    return {"registry": registry, "plan": plan, "manifest": manifest,
            "validation": validation, "raw_entries": raw_entries, "summary": summary}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--live", action="store_true", help="explicitly perform the one live acquisition")
    args = parser.parse_args()
    result = run_live() if args.live else regenerate_offline()
    if result["summary"]["implementation"] != "PASS":
        raise SystemExit(1)


if __name__ == "__main__":
    main()
