#!/usr/bin/env python3
"""Focused exact canonical/schema/digest tests; provider access is forbidden."""

import copy
import hashlib
import io
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill


sys.dont_write_bytecode = True
STEP_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = STEP_DIR.parent
ACQUISITION_DIR = STEP_DIR / "00_tool" / "acquisition"
for import_path in (PROJECT_ROOT, ACQUISITION_DIR, STEP_DIR / "00_tool"):
    if str(import_path) not in sys.path:
        sys.path.insert(0, str(import_path))

from common.json_utils import write_jsonl
from google_sheet_acquisition_contract import (
    INT64_MAX,
    INT64_MIN,
    MANIFEST_DIGEST_FIELDS,
    MANIFEST_DOMAIN,
    MANIFEST_FIELDS,
    ORDERED_SNAPSHOT_SET_DOMAIN,
    PLANNED_CONTAINER_SET_DIGEST_FIELDS,
    PLANNED_CONTAINER_SET_DOMAIN,
    PROFILE_DIGEST_FIELDS,
    PROFILE_DOMAIN,
    RESOLVED_SCOPE_DIGEST_FIELDS,
    RESOLVED_SCOPE_DOMAIN,
    SNAPSHOT_ENTRY_FIELDS,
    ContractError,
    calculate_entry_raw_digest,
    calculate_manifest_digest,
    calculate_ordered_snapshot_set_digest,
    calculate_planned_container_set_digest,
    calculate_profile_digest,
    calculate_resolved_scope_digest,
    canonical_json_bytes,
    finalize_manifest,
    finalize_profile,
    offline_negative_proofs,
    parse_canonical_json,
    validate_manifest,
    validate_snapshot_entry,
)
from run_google_sheet_acquisition_prototype import (
    DIGEST_SCHEMA_NEGATIVE_CASE_IDS,
    GOLDEN_PATH,
    RESULT_DIR,
    _build_snapshot_and_manifest,
    _load_profile_registry,
    _strict_implementation_pass,
    _unknown_source_version,
    build_attempt_plan,
    observe_workbook,
    select_representative_locator,
)


EXACT_TIME = "2026-08-27T12:34:56.000000Z"


def _direct_digest(domain, payload):
    return "sha256:" + hashlib.sha256(domain.encode("utf-8") + b"\x00" + payload).hexdigest()


def _projection(value, fields):
    return {field: copy.deepcopy(value[field]) for field in fields}


def _exception_negative_case(case_id, operation):
    expected = {"exception": "ContractError"}
    try:
        operation()
        actual = {"exception": "NONE"}
    except Exception as error:  # Evidence records an unexpected exception type too.
        actual = {
            "exception": type(error).__name__,
            "reason": getattr(error, "reason", str(error)),
        }
    return {
        "case_id": case_id,
        "expected": expected,
        "actual": actual,
        "passed": actual["exception"] == expected["exception"],
    }


def _snapshot_negative_case(case_id, entry, expected_reason):
    reasons = validate_snapshot_entry(entry, True)
    expected = {"rejected": True, "reason_contains": expected_reason}
    actual = {"rejected": bool(reasons), "reasons": reasons}
    return {
        "case_id": case_id,
        "expected": expected,
        "actual": actual,
        "passed": actual["rejected"] is True and any(
            expected_reason in reason for reason in reasons
        ),
    }


def _manifest_negative_case(
    case_id, manifest, registry, plan, raw_entries, expected_reason,
):
    result = validate_manifest(manifest, registry, plan, raw_entries)
    expected = {
        "valid": False,
        "acquisition_status": "INCOMPLETE",
        "eligible": 0,
        "reason_contains": expected_reason,
    }
    actual = {
        "valid": result["valid"],
        "acquisition_status": result["acquisition_status"],
        "eligible": result["eligible"],
        "reasons": result["reasons"],
    }
    return {
        "case_id": case_id,
        "expected": expected,
        "actual": actual,
        "passed": (
            actual["valid"] is False
            and actual["acquisition_status"] == "INCOMPLETE"
            and actual["eligible"] == 0
            and any(expected_reason in reason for reason in actual["reasons"])
        ),
    }


def digest_schema_negative_cases(manifest, registry, plan, raw_entries):
    entry = manifest["snapshot_entries"][0]
    cases = [
        _exception_negative_case(
            "nfc_key_collision",
            lambda: canonical_json_bytes({"e\u0301": 1, "é": 2}),
        ),
        _exception_negative_case(
            "duplicate_json_key", lambda: parse_canonical_json(b'{"a":1,"a":2}'),
        ),
        _exception_negative_case(
            "lone_surrogate", lambda: canonical_json_bytes("\ud800"),
        ),
    ]
    bool_entry = copy.deepcopy(entry)
    bool_entry["sequence"] = True
    cases.append(_snapshot_negative_case("bool_as_integer", bool_entry, "not_int64"))
    cases.extend([
        _exception_negative_case("float", lambda: canonical_json_bytes(1.0)),
        _exception_negative_case("exponent", lambda: parse_canonical_json(b"1e2")),
        _exception_negative_case("nan", lambda: parse_canonical_json(b"NaN")),
        _exception_negative_case("leading_plus", lambda: parse_canonical_json(b"+1")),
        _exception_negative_case("leading_zero", lambda: parse_canonical_json(b"01")),
        _exception_negative_case(
            "int64_max_plus_one", lambda: canonical_json_bytes(INT64_MAX + 1),
        ),
        _exception_negative_case(
            "int64_min_minus_one", lambda: canonical_json_bytes(INT64_MIN - 1),
        ),
    ])
    for case_id, invalid_datetime in (
        ("datetime_offset", "2026-08-27T12:34:56.000000+00:00"),
        ("datetime_fraction_missing", "2026-08-27T12:34:56Z"),
    ):
        changed = copy.deepcopy(entry)
        changed["retrieved_at"] = invalid_datetime
        cases.append(_snapshot_negative_case(case_id, changed, "datetime"))
    unknown_manifest = copy.deepcopy(manifest)
    unknown_manifest["unknown"] = "reject"
    cases.append(_manifest_negative_case(
        "manifest_unknown_field", unknown_manifest, registry, plan, raw_entries,
        "unknown_field",
    ))
    unknown_snapshot = copy.deepcopy(entry)
    unknown_snapshot["unknown"] = "reject"
    cases.append(_snapshot_negative_case(
        "snapshot_unknown_field", unknown_snapshot, "unknown_field",
    ))
    unknown_nested = copy.deepcopy(entry)
    unknown_nested["page_or_tab_id"]["unknown"] = "reject"
    cases.append(_snapshot_negative_case(
        "nested_unknown_field", unknown_nested, "unknown_field",
    ))
    null_status = copy.deepcopy(entry)
    null_status["http_status"] = None
    cases.append(_snapshot_negative_case(
        "http_status_null", null_status, "not_int64",
    ))
    missing_status = copy.deepcopy(entry)
    del missing_status["http_status"]
    cases.append(_snapshot_negative_case(
        "http_status_missing_for_http", missing_status, "missing_field:http_status",
    ))
    changed_raw_entries = dict(raw_entries)
    changed_raw_entries[entry["entry_id"]] = raw_entries[entry["entry_id"]] + b"-mutated"
    cases.append(_manifest_negative_case(
        "entry_raw_digest_mismatch", manifest, registry, plan, changed_raw_entries,
        "snapshot_entry_digest_mismatch",
    ))
    manifest_digest_mismatch = copy.deepcopy(manifest)
    manifest_digest_mismatch["manifest_digest"] = "sha256:" + "f" * 64
    cases.append(_manifest_negative_case(
        "manifest_digest_mismatch", manifest_digest_mismatch, registry, plan,
        raw_entries, "manifest_digest_mismatch",
    ))
    ordered_set_digest_mismatch = copy.deepcopy(manifest)
    ordered_set_digest_mismatch["ordered_snapshot_set_digest"] = "sha256:" + "f" * 64
    cases.append(_manifest_negative_case(
        "ordered_set_digest_mismatch", ordered_set_digest_mismatch, registry, plan,
        raw_entries, "ordered_snapshot_set_digest_mismatch",
    ))
    return cases


class GoogleSheetDigestConformanceTest(unittest.TestCase):
    def setUp(self):
        self.registry = _load_profile_registry()
        profile = copy.deepcopy(self.registry["profiles"][0])
        profile["presentation_policy"] = "NOT_USED"
        self.registry["profiles"][0] = finalize_profile(profile)
        self.selected = {
            "spreadsheet_key": "syntheticSheetKey",
            "gid": "123",
            "normalized_locator": "https://docs.google.com/spreadsheets/d/syntheticSheetKey/edit#gid=123",
            "evidence_message_ids": ["synthetic-message"],
            "evidence_count": 1,
            "link_text": "その他人材情報一覧",
        }
        self.plan = build_attempt_plan(self.registry, self.selected, EXACT_TIME)
        self.raw = b"offline-snapshot"
        observation = {
            "version_kind": "PROVIDER_VERSION",
            "version_scope": "WORKBOOK_WIDE",
            "version_strength": "STRONG",
            "version_binding_id": "revision:stable",
            "provider_authority_ref": "provider:google-sheets",
            "observed_at": EXACT_TIME,
        }
        authority = {**observation, "pre_version": copy.deepcopy(observation), "post_version": copy.deepcopy(observation)}
        self.manifest, self.raw_entries = _build_snapshot_and_manifest(
            self.registry, self.plan, self.raw, EXACT_TIME, authority, 200
        )

    def assertContractRejects(self, operation):
        with self.assertRaises(ContractError):
            operation()

    def test_01_fixed_golden_canonical_bytes_and_digest(self):
        golden = json.loads(GOLDEN_PATH.read_text(encoding="utf-8"))
        expected_bytes = golden["expected_canonical_json"].encode("utf-8")
        self.assertEqual(expected_bytes, canonical_json_bytes(golden["logical_object"]))
        self.assertFalse(expected_bytes.startswith(b"\xef\xbb\xbf"))
        self.assertFalse(expected_bytes.endswith(b"\n"))
        self.assertEqual(
            "sha256:8434afd8683e87ecda503124b0dc6a3c8ae2b3d38986ae9eb2e719fe0b40d443",
            calculate_ordered_snapshot_set_digest(golden["logical_object"]),
        )

    def test_02_control_escaping_and_string_rules(self):
        self.assertEqual(b'"\\u000a"', canonical_json_bytes("\n"))
        self.assertEqual(b'"\\u0009"', canonical_json_bytes("\t"))
        self.assertEqual(b'"\\u0000\\u001f"', canonical_json_bytes("\x00\x1f"))
        self.assertEqual(b'"\\\"\\\\/"', canonical_json_bytes('"\\/'))
        encoded = canonical_json_bytes("シートA")
        self.assertEqual('"シートA"'.encode("utf-8"), encoded)
        self.assertNotIn(b"\\u30", encoded)

    def test_03_nfc_key_order_collision_duplicate_and_surrogate(self):
        self.assertEqual(canonical_json_bytes("e\u0301"), canonical_json_bytes("é"))
        self.assertEqual('{"a":2,"é":1}'.encode("utf-8"), canonical_json_bytes({"e\u0301": 1, "a": 2}))
        self.assertContractRejects(lambda: canonical_json_bytes({"e\u0301": 1, "é": 2}))
        self.assertContractRejects(lambda: parse_canonical_json(b'{"a":1,"a":2}'))
        self.assertContractRejects(lambda: canonical_json_bytes("\ud800"))

    def test_04_integer_contract(self):
        self.assertEqual(str(INT64_MAX).encode(), canonical_json_bytes(INT64_MAX))
        self.assertEqual(str(INT64_MIN).encode(), canonical_json_bytes(INT64_MIN))
        self.assertContractRejects(lambda: canonical_json_bytes(INT64_MAX + 1))
        self.assertContractRejects(lambda: canonical_json_bytes(INT64_MIN - 1))
        self.assertContractRejects(lambda: canonical_json_bytes(1.0))
        self.assertContractRejects(lambda: parse_canonical_json(b"1e2"))
        self.assertContractRejects(lambda: parse_canonical_json(b"NaN"))
        self.assertContractRejects(lambda: parse_canonical_json(b"+1"))
        self.assertContractRejects(lambda: parse_canonical_json(b"01"))
        self.assertEqual(b"0", canonical_json_bytes(json.loads("-0")))
        changed = copy.deepcopy(self.manifest["snapshot_entries"][0])
        changed["sequence"] = True
        self.assertTrue(any("not_int64" in reason for reason in validate_snapshot_entry(changed, True)))

    def test_05_datetime_exact_and_negative(self):
        entry = self.manifest["snapshot_entries"][0]
        self.assertEqual([], validate_snapshot_entry(entry, True))
        for invalid in (
            "2026-08-27T12:34:56+00:00",
            "2026-08-27T12:34:56Z",
            "2026-08-27T12:34:60.000000Z",
        ):
            changed = copy.deepcopy(entry)
            changed["retrieved_at"] = invalid
            self.assertTrue(any("datetime" in reason for reason in validate_snapshot_entry(changed, True)))

    def test_06_exact_manifest_and_snapshot_schema(self):
        self.assertEqual(15, len(self.manifest))
        self.assertEqual(set(MANIFEST_FIELDS), set(self.manifest))
        entry = self.manifest["snapshot_entries"][0]
        self.assertEqual(18, len(entry))
        self.assertEqual(set(SNAPSHOT_ENTRY_FIELDS), set(entry))
        changed_manifest = copy.deepcopy(self.manifest)
        changed_manifest["unknown"] = "reject"
        result = validate_manifest(changed_manifest, self.registry, self.plan, self.raw_entries)
        self.assertFalse(result["exact_manifest_schema"])
        changed_entry = copy.deepcopy(entry)
        changed_entry["unknown"] = "reject"
        self.assertTrue(any("unknown_field" in reason for reason in validate_snapshot_entry(changed_entry, True)))
        nested_entry = copy.deepcopy(entry)
        nested_entry["page_or_tab_id"]["unknown"] = "reject"
        self.assertTrue(any("unknown_field" in reason for reason in validate_snapshot_entry(nested_entry, True)))
        null_entry = copy.deepcopy(entry)
        null_entry["http_status"] = None
        self.assertTrue(validate_snapshot_entry(null_entry, True))
        missing_entry = copy.deepcopy(entry)
        del missing_entry["http_status"]
        self.assertFalse(validate_snapshot_entry(missing_entry, False))
        self.assertTrue(validate_snapshot_entry(missing_entry, True))

    def test_07_entry_raw_digest_domain_is_independent_oracle(self):
        expected = _direct_digest("99-1/source-acquisition/entry-raw/v1", self.raw)
        self.assertEqual(expected, calculate_entry_raw_digest(self.raw))
        self.assertNotEqual("sha256:" + hashlib.sha256(self.raw).hexdigest(), expected)

    def test_08_ordered_set_projection_and_sequence_sensitivity(self):
        entry = self.manifest["snapshot_entries"][0]
        direct_projection = [_projection(entry, (
            "sequence", "entry_id", "planned_container_id", "page_or_tab_id",
            "locator_binding_id", "requested_range", "returned_range", "source_version",
            "content_type", "byte_length", "entry_raw_digest", "next_locator_metadata",
            "terminal", "terminal_evidence",
        ))]
        expected = _direct_digest(ORDERED_SNAPSHOT_SET_DOMAIN, canonical_json_bytes(direct_projection))
        self.assertEqual(expected, calculate_ordered_snapshot_set_digest([entry]))
        excluded_change = copy.deepcopy(entry)
        excluded_change["retrieved_at"] = "2026-08-27T12:34:57.000000Z"
        excluded_change["http_status"] = 206
        excluded_change["raw_artifact_ref"] = "snapshot/other.xlsx"
        self.assertEqual(expected, calculate_ordered_snapshot_set_digest([excluded_change]))
        included_change = copy.deepcopy(entry)
        included_change["sequence"] = 1
        self.assertNotEqual(expected, calculate_ordered_snapshot_set_digest([included_change]))

    def test_09_manifest_profile_scope_and_container_domains(self):
        manifest_expected = _direct_digest(
            MANIFEST_DOMAIN,
            canonical_json_bytes(_projection(self.manifest, MANIFEST_DIGEST_FIELDS)),
        )
        self.assertEqual(manifest_expected, calculate_manifest_digest(self.manifest))
        profile = self.registry["profiles"][0]
        self.assertEqual(
            _direct_digest(PROFILE_DOMAIN, canonical_json_bytes(_projection(profile, PROFILE_DIGEST_FIELDS))),
            calculate_profile_digest(profile),
        )
        scope = self.plan["resolved_planned_scope"]
        self.assertEqual(
            _direct_digest(RESOLVED_SCOPE_DOMAIN, canonical_json_bytes(_projection(scope, RESOLVED_SCOPE_DIGEST_FIELDS))),
            calculate_resolved_scope_digest(scope),
        )
        container_set = self.plan["planned_container_set"]
        direct_set = _projection(container_set, PLANNED_CONTAINER_SET_DIGEST_FIELDS)
        direct_set["planned_container_entries"] = [
            _projection(container_set["planned_container_entries"][0], (
                "sequence", "planned_container_id", "container_kind", "required",
                "logical_role", "locator_ref", "locator_binding_id",
                "acquisition_profile_id", "acquisition_profile_version",
                "acquisition_profile_digest", "optional_absence_rule",
            ))
        ]
        self.assertEqual(
            _direct_digest(PLANNED_CONTAINER_SET_DOMAIN, canonical_json_bytes(direct_set)),
            calculate_planned_container_set_digest(container_set),
        )

    def test_10_manifest_digest_includes_nested_and_optional_http_status(self):
        changed = copy.deepcopy(self.manifest)
        changed["snapshot_entries"][0]["retrieved_at"] = "2026-08-27T12:34:57.000000Z"
        self.assertNotEqual(calculate_manifest_digest(self.manifest), calculate_manifest_digest(changed))
        changed = copy.deepcopy(self.manifest)
        changed["snapshot_entries"][0]["http_status"] = 206
        self.assertNotEqual(calculate_manifest_digest(self.manifest), calculate_manifest_digest(changed))
        changed = copy.deepcopy(self.manifest)
        changed["manifest_digest"] = "sha256:" + "f" * 64
        self.assertEqual(calculate_manifest_digest(self.manifest), calculate_manifest_digest(changed))

    def test_11_valid_manifest_and_strict_fail_on_schema_or_digest(self):
        result = validate_manifest(self.manifest, self.registry, self.plan, self.raw_entries)
        self.assertTrue(result["valid"])
        self.assertTrue(result["exact_manifest_schema"])
        self.assertTrue(result["exact_snapshot_entry_schema"])
        self.assertTrue(result["digest_conformance"])
        self.assertEqual("VERIFIED_COMPLETE", result["acquisition_status"])
        changed = copy.deepcopy(self.manifest)
        changed["manifest_digest"] = "sha256:" + "f" * 64
        result = validate_manifest(changed, self.registry, self.plan, self.raw_entries)
        self.assertFalse(result["valid"])
        self.assertEqual("INCOMPLETE", result["acquisition_status"])
        individual_cases = digest_schema_negative_cases(
            self.manifest, self.registry, self.plan, self.raw_entries
        )
        self.assertEqual(
            list(DIGEST_SCHEMA_NEGATIVE_CASE_IDS),
            [case["case_id"] for case in individual_cases],
        )
        for case in individual_cases:
            with self.subTest(digest_schema_negative=case["case_id"]):
                self.assertTrue(case["passed"], case)
        focused = {
            "focused_status": "PASS",
            "focused_passed": 14,
            "focused_total": 14,
            "digest_schema_negative_passed": sum(
                case["passed"] is True for case in individual_cases
            ),
            "digest_schema_negative_total": len(individual_cases),
            "digest_schema_negative_cases": individual_cases,
            "false_pass_prevention": "PASS",
        }
        conformance = [
            {"name": "golden_canonical_bytes", "passed": True},
            {"name": "golden_ordered_set_digest", "passed": True},
        ]
        valid_result = validate_manifest(
            self.manifest, self.registry, self.plan, self.raw_entries
        )
        self.assertTrue(_strict_implementation_pass(
            valid_result, conformance, 9, 9, focused
        ))
        failed_focused = copy.deepcopy(focused)
        failed_focused["digest_schema_negative_cases"][0]["passed"] = False
        self.assertFalse(_strict_implementation_pass(
            valid_result, conformance, 9, 9, failed_focused
        ))

    def test_12_original_nine_negative_proofs(self):
        cases = offline_negative_proofs(self.manifest, self.registry, self.plan, self.raw_entries)
        expected = {
            "profile_digest_mismatch": ("UNVERIFIED", "profile_digest_mismatch"),
            "planned_scope_mismatch": ("PARTIAL", "planned_scope_mismatch"),
            "required_container_missing": ("PARTIAL", "required_container_missing"),
            "strong_version_unavailable": ("UNVERIFIED", "strong_version_unavailable"),
            "revision_drift": ("SNAPSHOT_UNSTABLE", "revision_drift"),
            "range_gap": ("PARTIAL", "range_gap"),
            "digest_mismatch": ("INCOMPLETE", "snapshot_entry_digest_mismatch"),
            "presentation_unresolved": ("UNVERIFIED", "presentation_unresolved"),
            "attempt_uncommitted": ("INCOMPLETE", "attempt_uncommitted"),
        }
        self.assertEqual(set(expected), {case["name"] for case in cases})
        for case in cases:
            status, reason = expected[case["name"]]
            with self.subTest(case=case["name"]):
                self.assertEqual(status, case["result"]["acquisition_status"])
                self.assertIn(reason, case["result"]["reasons"])
                self.assertEqual(0, case["result"]["eligible"])
        presentation = next(case for case in cases if case["name"] == "presentation_unresolved")
        self.assertEqual("HUMAN_REVIEW", presentation["result"]["review_status"])

    def test_13_workbook_metadata_observation_remains_unchanged(self):
        workbook = Workbook()
        active = workbook.active
        active.title = "人材一覧"
        active["A1"] = "氏名"
        active["A2"] = "匿名A"
        active["B2"] = "=1+1"
        active["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
        active.row_dimensions[2].hidden = True
        active.column_dimensions["C"].hidden = True
        active.conditional_formatting.add(
            "A1:A2",
            CellIsRule(operator="equal", formula=['"匿名A"'], fill=PatternFill(fill_type="solid", fgColor="00FF00")),
        )
        hidden = workbook.create_sheet("設定")
        hidden.sheet_state = "hidden"
        hidden["A1"] = "support"
        output = io.BytesIO()
        workbook.save(output)
        observed = observe_workbook(output.getvalue())
        self.assertEqual("AVAILABLE", observed["workbook_tab_inventory"])
        self.assertEqual(["VISIBLE", "HIDDEN"], [tab["visibility"] for tab in observed["tabs"]])
        self.assertEqual("PARTIAL", observed["presentation_metadata"]["overall_availability"])
        self.assertEqual("UNAVAILABLE", observed["presentation_metadata"]["effective_format"])

    def test_14_saved_gmail_selection_still_normalizes_one_locator(self):
        records = [{
            "message_id": "m1", "from": "Sales <sales@tanapism.co.jp>",
            "html_links": [{
                "text": "★☆その他人材情報一覧☆★",
                "href": "https://docs.google.com/spreadsheets/d/key123/edit?tracking=discard#gid=456",
            }],
        }]
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "gmail.jsonl"
            write_jsonl(str(path), records)
            selected = select_representative_locator(path, self.registry["profiles"][0])
        self.assertEqual("https://docs.google.com/spreadsheets/d/key123/edit#gid=456", selected["normalized_locator"])


def main():
    suite = unittest.defaultTestLoader.loadTestsFromTestCase(GoogleSheetDigestConformanceTest)
    result = unittest.TextTestRunner(verbosity=2).run(suite)
    failures = len(result.failures) + len(result.errors)
    evidence_fixture = GoogleSheetDigestConformanceTest(methodName="test_01_fixed_golden_canonical_bytes_and_digest")
    evidence_fixture.setUp()
    individual_cases = digest_schema_negative_cases(
        evidence_fixture.manifest,
        evidence_fixture.registry,
        evidence_fixture.plan,
        evidence_fixture.raw_entries,
    )
    individual_passed = sum(case["passed"] is True for case in individual_cases)
    evidence = {
        "evidence_schema_version": "GoogleSheetDigestFocusedTestEvidence.v1",
        "focused_passed": result.testsRun - failures,
        "focused_total": result.testsRun,
        "focused_status": "PASS" if result.wasSuccessful() else "FAIL",
        "digest_schema_negative_passed": individual_passed,
        "digest_schema_negative_total": len(individual_cases),
        "digest_schema_negative_cases": individual_cases,
        "google_live_access": 0,
        "production_write": 0,
        "pipeline_04_05_runs": 0,
    }
    valid_result = validate_manifest(
        evidence_fixture.manifest,
        evidence_fixture.registry,
        evidence_fixture.plan,
        evidence_fixture.raw_entries,
    )
    conformance = [
        {"name": "golden_canonical_bytes", "passed": True},
        {"name": "golden_ordered_set_digest", "passed": True},
    ]
    evidence["false_pass_prevention"] = "PASS"
    positive_pass = _strict_implementation_pass(valid_result, conformance, 9, 9, evidence)
    failed_evidence = copy.deepcopy(evidence)
    failed_evidence["digest_schema_negative_cases"][0]["passed"] = False
    false_pass_rejected = not _strict_implementation_pass(
        valid_result, conformance, 9, 9, failed_evidence
    )
    evidence["false_pass_prevention"] = (
        "PASS" if positive_pass and false_pass_rejected else "FAIL"
    )
    evidence["false_pass_check"] = {
        "mutated_case_id": failed_evidence["digest_schema_negative_cases"][0]["case_id"],
        "mutated_case_passed": False,
        "strict_pass": not false_pass_rejected,
        "passed": positive_pass and false_pass_rejected,
    }
    write_jsonl(str(RESULT_DIR / "focused_test_result.jsonl"), [evidence])
    if not result.wasSuccessful() or evidence["false_pass_prevention"] != "PASS":
        raise SystemExit(1)


if __name__ == "__main__":
    main()
